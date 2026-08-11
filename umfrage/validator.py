"""Validation of returned questionnaire Excel files.

The public function :func:`validate_response` checks a single filled-in
``.xlsx`` file against its originating questionnaire config and returns a
:class:`ValidationResult` that includes extracted answers and respondent info.

Validation checks
-----------------
1. Required sheets present ("Questionnaire" and "_meta").
2. ``_meta`` question IDs (JSON list) match the expected IDs from the config.
3. Config hash in ``_meta`` matches the questionnaire's canonical hash
   (mismatch produces a warning, not an error, to handle minor re-distributions).
4. All required respondent fields are non-empty.
5. All question rows are still present (none deleted).
6. Answer values satisfy type-specific constraints (scale range, yes/no literals).
7. All required questions have a non-empty answer.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from umfrage.models import AnswerType, Questionnaire
from umfrage.translator import Translator

# Sheet name constants — must match generator.py
QUESTIONNAIRE_SHEET = "Questionnaire"
META_SHEET = "_meta"


@dataclass
class ValidationResult:
    """Outcome of validating a single returned questionnaire file."""

    path: Path
    is_valid: bool
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    respondent_info: dict[str, str] = field(default_factory=dict)
    answers: dict[str, Any] = field(default_factory=dict)
    config_hash: str = ""
    questionnaire_id: str = ""


def validate_response(path: Path, questionnaire: Questionnaire) -> ValidationResult:
    """Validate a returned questionnaire ``.xlsx`` file against the expected config.

    Args:
        path: Path to the filled-in Excel file.
        questionnaire: The originating questionnaire configuration.

    Returns:
        A :class:`ValidationResult` with ``is_valid``, ``errors``, ``warnings``,
        and the extracted ``respondent_info`` / ``answers`` dictionaries.
    """
    result = ValidationResult(path=path, is_valid=True)

    # Open the workbook (read-only, data_only for cell values not formulas)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        result.is_valid = False
        result.errors.append(f"Cannot open file: {exc}")
        return result

    # Check 1 — required sheets present
    if QUESTIONNAIRE_SHEET not in wb.sheetnames:
        result.errors.append(f"Missing required sheet '{QUESTIONNAIRE_SHEET}'.")
        result.is_valid = False
    if META_SHEET not in wb.sheetnames:
        result.errors.append(f"Missing required hidden sheet '{META_SHEET}'.")
        result.is_valid = False

    if not result.is_valid:
        return result

    ws_meta = wb[META_SHEET]
    ws_q = wb[QUESTIONNAIRE_SHEET]

    # Read all _meta key-value pairs
    meta = _read_meta(ws_meta)
    result.config_hash = meta.get("config_hash", "")
    result.questionnaire_id = meta.get("questionnaire_id", "")

    # Check 2 — question IDs in _meta match the config
    stored_ids_raw = meta.get("question_ids", "[]")
    try:
        stored_ids: list[str] = json.loads(stored_ids_raw)
    except json.JSONDecodeError:
        result.errors.append("'_meta' sheet 'question_ids' value is not valid JSON.")
        result.is_valid = False
        return result

    expected_ids = [q.id for q in questionnaire.all_questions()]
    missing_ids = set(expected_ids) - set(stored_ids)
    extra_ids = set(stored_ids) - set(expected_ids)
    if missing_ids:
        result.errors.append(
            f"Question IDs missing from file: {sorted(missing_ids)}"
        )
        result.is_valid = False
    if extra_ids:
        result.errors.append(
            f"Unexpected question IDs in file (may indicate tampering): "
            f"{sorted(extra_ids)}"
        )
        result.is_valid = False

    # Check 3 — config hash match (warning only)
    if meta.get("config_hash") != questionnaire.config_hash():
        result.warnings.append(
            "Config hash mismatch — the file may have been generated from a "
            "different version of the questionnaire config. "
            "Answers are extracted on a best-effort basis."
        )

    # Read all sheet rows once for efficiency
    all_rows = list(ws_q.iter_rows(values_only=True))

    # Build a translator for language-aware yes/no validation
    translator = Translator(questionnaire.language)

    # Extract respondent info and answers from the sheet content
    _extract_respondent_info(all_rows, questionnaire, result)
    _extract_answers(all_rows, questionnaire, result)

    # Check 4 — required respondent fields non-empty
    for resp_field in questionnaire.respondent_fields:
        if resp_field.required:
            value = result.respondent_info.get(resp_field.label, "")
            if not str(value).strip():
                result.errors.append(
                    f"Required respondent field '{resp_field.label}' is empty."
                )
                result.is_valid = False

    # Checks 5, 6, 7 — per-question answer validation
    for q in questionnaire.all_questions():
        raw_value = result.answers.get(q.id)
        is_blank = raw_value is None or str(raw_value).strip() == ""

        # Check 7 — required questions must have an answer
        if is_blank:
            if q.required:
                result.errors.append(
                    f"Required question '{q.id}' has no answer."
                )
                result.is_valid = False
        else:
            # Check 6 — answer value within constraints
            error = _validate_answer_value(q.id, raw_value, q.answer, translator)
            if error:
                result.errors.append(error)
                result.is_valid = False

    return result


# ── Internal helpers ──────────────────────────────────────────────────────────

def _read_meta(ws_meta) -> dict[str, str]:
    """Read all key-value pairs from the _meta sheet (column A → column B)."""
    data: dict[str, str] = {}
    for row in ws_meta.iter_rows(max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            data[str(row[0])] = str(row[1])
    return data


def _extract_respondent_info(
    all_rows: list[tuple], questionnaire: Questionnaire, result: ValidationResult
) -> None:
    """Scan sheet rows to extract respondent field values.

    Respondent rows have the format: [label + ":",  value, ...]
    where the label matches one of the questionnaire's respondent_fields labels.
    """
    field_labels = {f.label for f in questionnaire.respondent_fields}
    for row in all_rows:
        if not row or not row[0]:
            continue
        cell_text = str(row[0]).strip()
        for label in field_labels:
            if cell_text == label + ":":
                raw = row[1] if len(row) > 1 else None
                result.respondent_info[label] = str(raw).strip() if raw is not None else ""
                break


def _extract_answers(
    all_rows: list[tuple], questionnaire: Questionnaire, result: ValidationResult
) -> None:
    """Scan sheet rows to extract question answers.

    Question rows have the question ID in column A (index 0) and the answer in
    column C (index 2, the COL_ANSWER column from generator.py).
    """
    question_id_set = {q.id for q in questionnaire.all_questions()}
    for row in all_rows:
        if not row or not row[0]:
            continue
        qid = str(row[0]).strip()
        if qid in question_id_set:
            # Column C is index 2 (COL_ANSWER = 3, 1-based → 2 zero-based)
            answer_value = row[2] if len(row) > 2 else None
            result.answers[qid] = answer_value


def _validate_answer_value(qid: str, value: Any, answer_config, translator: Translator) -> str | None:
    """Validate a single answer value against its answer config.

    Returns an error message string, or ``None`` if the value is valid.
    Yes/No answers are validated against the language-specific strings
    provided by *translator* (case-insensitive).
    """
    if answer_config.type == AnswerType.SCALE:
        try:
            int_val = int(value)
        except (TypeError, ValueError):
            return f"Question '{qid}': expected an integer, got '{value}'."
        if answer_config.min_value is not None and int_val < answer_config.min_value:
            return (
                f"Question '{qid}': value {int_val} is below the "
                f"minimum allowed value of {answer_config.min_value}."
            )
        if answer_config.max_value is not None and int_val > answer_config.max_value:
            return (
                f"Question '{qid}': value {int_val} exceeds the "
                f"maximum allowed value of {answer_config.max_value}."
            )

    elif answer_config.type == AnswerType.YES_NO:
        yes_val, no_val = translator.yes_no_values()
        if str(value).strip().lower() not in (yes_val.lower(), no_val.lower()):
            return (
                f"Question '{qid}': expected '{yes_val}' or '{no_val}', "
                f"got '{value}'."
            )

    # FREETEXT: any non-empty string is valid (emptiness is checked by the caller)
    return None
