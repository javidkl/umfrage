"""Helpers that convert StyleConfig values into openpyxl formatting objects.

Each ``apply_*`` function mutates a single cell in place. All functions accept
the full :class:`~umfrage.models.StyleConfig` so that colour decisions can be
made consistently from one place.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from umfrage.models import CellStyle, StyleConfig


# ── Low-level factory helpers ─────────────────────────────────────────────────

def make_font(style: CellStyle, size_override: int | None = None) -> Font:
    """Build an :class:`openpyxl.styles.Font` from a :class:`~umfrage.models.CellStyle`."""
    return Font(
        bold=style.bold,
        italic=style.italic,
        color=style.font_color,
        size=size_override if size_override is not None else style.font_size,
    )


def make_fill(hex_color: str) -> PatternFill:
    """Build a solid :class:`openpyxl.styles.PatternFill` from a hex color string."""
    return PatternFill(fill_type="solid", fgColor=hex_color)


def make_thin_border() -> Border:
    """Return a thin all-around border suitable for data cells."""
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def make_medium_border() -> Border:
    """Return a slightly heavier border used for editable answer cells."""
    medium = Side(style="medium", color="888888")
    return Border(left=medium, right=medium, top=medium, bottom=medium)


# ── Cell-category applicators ─────────────────────────────────────────────────

def apply_header_style(cell, style: StyleConfig) -> None:
    """Apply the large title/header row style to *cell*."""
    s = style.header
    cell.font = make_font(s)
    cell.fill = make_fill(s.background_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.protection = Protection(locked=True)


def apply_section_style(cell, style: StyleConfig) -> None:
    """Apply the section-header row style to *cell*."""
    s = style.section_header
    cell.font = make_font(s)
    cell.fill = make_fill(s.background_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = make_thin_border()
    cell.protection = Protection(locked=True)


def apply_question_style(cell, style: StyleConfig, alternate: bool = False) -> None:
    """Apply question-row style to *cell* (locked; alternates background if requested)."""
    s = style.question_row
    color = s.alternate_color if (alternate and s.alternate_color) else s.background_color
    cell.font = make_font(s)
    cell.fill = make_fill(color)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = make_thin_border()
    cell.protection = Protection(locked=True)


def apply_answer_style(cell, style: StyleConfig) -> None:
    """Apply answer-cell style to *cell* (explicitly **unlocked** and editable)."""
    s = style.answer_cell
    cell.font = make_font(s)
    cell.fill = make_fill(s.background_color)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = make_medium_border()
    cell.protection = Protection(locked=False)


def apply_respondent_header_style(cell, style: StyleConfig) -> None:
    """Apply the 'RESPONDENT INFORMATION' section header style to *cell*."""
    s = style.respondent_header
    cell.font = make_font(s)
    cell.fill = make_fill(s.background_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.protection = Protection(locked=True)


def apply_result_header_style(cell, style: StyleConfig) -> None:
    """Apply the result-sheet column header style to *cell*."""
    s = style.result_header
    cell.font = make_font(s)
    cell.fill = make_fill(s.background_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_footer_style(cell) -> None:
    """Apply subtle footer row style to *cell* (small, italic, gray, locked)."""
    cell.font = Font(bold=False, italic=True, color="888888", size=8)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.protection = Protection(locked=True)
    cell.border = make_thin_border()
