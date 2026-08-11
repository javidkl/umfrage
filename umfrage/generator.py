"""Excel questionnaire generator for the umfrage tool.

The main public function is :func:`generate_questionnaire`, which produces a
single ``.xlsx`` file from a validated :class:`~umfrage.models.Questionnaire`
and :class:`~umfrage.models.StyleConfig`.

Layout of the generated "Questionnaire" sheet
----------------------------------------------
Row 1   : Title (merged, styled as header)
Row 2   : Organizer info (merged, styled as sub-header)
Row 3   : Spacer
Row 4   : "RESPONDENT INFORMATION" label (merged, locked)
Row 5+  : One row per respondent field (label locked, value cell editable)
Row N   : Spacer
Row N+1 : Section header (merged, locked)
Row N+2 : Column headers: ID | Question | Answer | Scale/Comment (locked)
Row N+3+: One row per question (ID, text, comment locked; answer cell editable)
           ... repeated for each section

Hidden "_meta" sheet
--------------------
Stores structural metadata for later validation:
  questionnaire_id, config_hash, version, generated (ISO timestamp),
  title, question_ids (JSON array), respondent_fields (JSON array),
  questionnaire_json (full serialized model for metadata-file reconstruction).
"""

from __future__ import annotations

import datetime
import json
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from umfrage.models import AnswerType, Questionnaire, StyleConfig
from umfrage.styles import (
    apply_answer_style,
    apply_header_style,
    apply_question_style,
    apply_respondent_header_style,
    apply_section_style,
    make_fill,
    make_font,
)
from umfrage.translator import Translator

# Column indices (1-based) for the Questionnaire sheet
COL_ID = 1       # A — question identifier
COL_TEXT = 2     # B — question text
COL_ANSWER = 3   # C — answer input (editable)
COL_COMMENT = 4  # D — scale labels / comments
TOTAL_COLS = 4


def generate_questionnaire(
    questionnaire: Questionnaire,
    style: StyleConfig,
    output_path: Path,
) -> Path:
    """Generate a protected Excel questionnaire file.

    Creates two sheets:

    * **"Questionnaire"** – the form to be filled in by respondents.
    * **"_meta"** – hidden sheet with structural metadata used during collection.

    Worksheet protection is applied so that only explicitly unlocked answer
    cells and respondent-field cells can be edited. An optional password can
    be configured via :attr:`~umfrage.models.StyleConfig.protection_password`.

    Args:
        questionnaire: Validated questionnaire model.
        style: Excel styling and protection configuration.
        output_path: Destination path for the ``.xlsx`` file.

    Returns:
        *output_path* after the file has been written.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionnaire"

    translator = Translator(questionnaire.language)
    _build_questionnaire_sheet(ws, questionnaire, style, translator)
    _build_meta_sheet(wb, questionnaire)
    _apply_sheet_protection(ws, style)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_metadata_file(questionnaire: Questionnaire, output_path: Path) -> Path:
    """Write a companion ``*_metadata.yaml`` file alongside a generated questionnaire.

    The file embeds the **full serialized Questionnaire model** so that
    ``umfrage collect`` can reconstruct the config without the original
    ``questionnaire.yaml`` being present.

    Args:
        questionnaire: The questionnaire to serialize.
        output_path: Destination path for the metadata YAML file.

    Returns:
        *output_path* after the file has been written.
    """
    metadata = {
        "questionnaire_id": questionnaire.questionnaire_id(),
        "config_hash": questionnaire.config_hash(),
        "generation_timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "title": questionnaire.title,
        "version": questionnaire.version,
        "question_ids": [q.id for q in questionnaire.all_questions()],
        "respondent_fields": [f.label for f in questionnaire.respondent_fields],
        "questionnaire": json.loads(questionnaire.model_dump_json()),
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        yaml.dump(metadata, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    return output_path


# ── Internal sheet builders ───────────────────────────────────────────────────

def _build_questionnaire_sheet(ws, questionnaire: Questionnaire, style: StyleConfig, translator: Translator) -> None:
    """Populate the main 'Questionnaire' worksheet."""
    cw = style.column_widths
    ws.column_dimensions[get_column_letter(COL_ID)].width = cw.question_id
    ws.column_dimensions[get_column_letter(COL_TEXT)].width = cw.question_text
    ws.column_dimensions[get_column_letter(COL_ANSWER)].width = cw.answer
    ws.column_dimensions[get_column_letter(COL_COMMENT)].width = cw.comment

    row = 1

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 32
    title_cell = ws.cell(row=row, column=COL_ID, value=questionnaire.title)
    apply_header_style(title_cell, style)
    ws.merge_cells(
        start_row=row, start_column=COL_ID, end_row=row, end_column=TOTAL_COLS
    )
    row += 1

    # ── Row 2: Organizer info ─────────────────────────────────────────────────
    org = questionnaire.organizer
    organizer_text = (
        f"{translator.t('label_organizer')}: {org.name}  |  "
        f"{translator.t('label_institution')}: {org.institution}  |  "
        f"{translator.t('label_email')}: {org.email}"
        + (f"  |  {translator.t('label_phone')}: {org.phone}" if org.phone else "")
    )
    ws.row_dimensions[row].height = 18
    org_cell = ws.cell(row=row, column=COL_ID, value=organizer_text)
    org_cell.font = make_font(style.header, size_override=9)
    org_cell.fill = make_fill(style.header.background_color)
    org_cell.alignment = Alignment(horizontal="left", vertical="center")
    org_cell.protection = Protection(locked=True)
    ws.merge_cells(
        start_row=row, start_column=COL_ID, end_row=row, end_column=TOTAL_COLS
    )
    row += 1

    # ── Row 3: Spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 8
    row += 1

    # ── Respondent Information section ───────────────────────────────────────
    ws.row_dimensions[row].height = 18
    resp_label = ws.cell(row=row, column=COL_ID, value=translator.t("label_respondent_information"))
    apply_respondent_header_style(resp_label, style)
    ws.merge_cells(
        start_row=row, start_column=COL_ID, end_row=row, end_column=TOTAL_COLS
    )
    row += 1

    for resp_field in questionnaire.respondent_fields:
        ws.row_dimensions[row].height = 20
        label_cell = ws.cell(row=row, column=COL_ID, value=resp_field.label + ":")
        label_cell.font = make_font(style.question_row, size_override=10)
        label_cell.fill = make_fill(style.question_row.background_color)
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.protection = Protection(locked=True)

        value_cell = ws.cell(row=row, column=COL_TEXT, value="")
        apply_answer_style(value_cell, style)
        ws.merge_cells(
            start_row=row, start_column=COL_TEXT, end_row=row, end_column=TOTAL_COLS
        )
        row += 1

    # ── Spacer ────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 8
    row += 1

    # ── Sections and questions ────────────────────────────────────────────────
    question_index = 0
    for section in questionnaire.sections:
        # Section header row
        ws.row_dimensions[row].height = 20
        section_cell = ws.cell(row=row, column=COL_ID, value=f"  {section.title.upper()}")
        apply_section_style(section_cell, style)
        ws.merge_cells(
            start_row=row, start_column=COL_ID, end_row=row, end_column=TOTAL_COLS
        )
        row += 1

        # Column header row
        ws.row_dimensions[row].height = 14
        for col, header in zip(
            [COL_ID, COL_TEXT, COL_ANSWER, COL_COMMENT],
            [
                translator.t("col_id"),
                translator.t("col_question"),
                translator.t("col_answer"),
                translator.t("col_scale_comment"),
            ],
        ):
            cell = ws.cell(row=row, column=col, value=header)
            apply_section_style(cell, style)
            cell.font = make_font(style.section_header, size_override=9)
        row += 1

        # Question rows
        for q in section.questions:
            alternate = question_index % 2 == 1
            ws.row_dimensions[row].height = 32

            id_cell = ws.cell(row=row, column=COL_ID, value=q.id)
            apply_question_style(id_cell, style, alternate)

            text_cell = ws.cell(row=row, column=COL_TEXT, value=q.text)
            apply_question_style(text_cell, style, alternate)
            text_cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

            answer_cell = ws.cell(row=row, column=COL_ANSWER, value="")
            apply_answer_style(answer_cell, style)
            _add_data_validation(ws, answer_cell, q.answer, translator)

            comment_text = _build_comment_text(q, translator)
            comment_cell = ws.cell(row=row, column=COL_COMMENT, value=comment_text)
            apply_question_style(comment_cell, style, alternate)
            comment_cell.font = make_font(style.question_row, size_override=8)
            comment_cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

            question_index += 1
            row += 1


def _build_comment_text(q, translator: Translator) -> str:
    """Compose the Scale/Comment column text for a question."""
    if q.comment:
        return q.comment
    ans = q.answer
    if ans.type == AnswerType.SCALE and ans.min_value is not None:
        if ans.description:
            return f"[{ans.min_value}–{ans.max_value}]  {ans.description}"
        return translator.t("hint_scale", min=ans.min_value, max=ans.max_value)
    if ans.type == AnswerType.YES_NO:
        return ans.description or translator.t("hint_yesno")
    if ans.type == AnswerType.FREETEXT:
        return ans.description or translator.t("hint_freetext")
    return ""


def _add_data_validation(ws, cell, answer_config, translator: Translator) -> None:
    """Attach Excel data validation to an answer cell based on the answer type."""
    if answer_config.type == AnswerType.SCALE:
        if answer_config.min_value is not None and answer_config.max_value is not None:
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1=str(answer_config.min_value),
                formula2=str(answer_config.max_value),
                showErrorMessage=True,
                errorTitle=translator.t("dv_error_title"),
                error=translator.t(
                    "dv_scale_error",
                    min=answer_config.min_value,
                    max=answer_config.max_value,
                ),
            )
            ws.add_data_validation(dv)
            dv.add(cell)

    elif answer_config.type == AnswerType.YES_NO:
        yes_val, no_val = translator.yes_no_values()
        dv = DataValidation(
            type="list",
            formula1=translator.yes_no_formula(),
            showErrorMessage=True,
            errorTitle=translator.t("dv_error_title"),
            error=translator.t("dv_yesno_error", yes=yes_val, no=no_val),
        )
        ws.add_data_validation(dv)
        dv.add(cell)
    # FREETEXT: no data validation added; any text is accepted.


def _build_meta_sheet(wb: Workbook, questionnaire: Questionnaire) -> None:
    """Create the hidden '_meta' sheet with structural metadata."""
    ws_meta = wb.create_sheet(title="_meta")
    ws_meta.sheet_state = "hidden"

    rows = [
        ("questionnaire_id", questionnaire.questionnaire_id()),
        ("config_hash", questionnaire.config_hash()),
        ("version", questionnaire.version),
        ("generated", datetime.datetime.now(datetime.timezone.utc).isoformat()),
        ("title", questionnaire.title),
        ("question_ids", json.dumps([q.id for q in questionnaire.all_questions()])),
        ("respondent_fields", json.dumps([f.label for f in questionnaire.respondent_fields])),
        # Full model embedded so the collector can reconstruct config without the
        # original questionnaire.yaml being available.
        ("questionnaire_json", questionnaire.model_dump_json()),
    ]
    for row_idx, (key, value) in enumerate(rows, start=1):
        ws_meta.cell(row=row_idx, column=1, value=key)
        ws_meta.cell(row=row_idx, column=2, value=value)


def _apply_sheet_protection(ws, style: StyleConfig) -> None:
    """Enable worksheet protection with an optional password."""
    ws.protection.sheet = True
    # Allow selecting locked and unlocked cells so respondents can navigate freely.
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    if style.protection_password:
        ws.protection.set_password(style.protection_password)
