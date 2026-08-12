"""Collection and aggregation of returned questionnaire Excel files.

Public API
----------
:func:`collect_all`
    Top-level entry point. Scans a folder, groups response files by their
    embedded questionnaire identity (config hash), resolves the config for
    each group, and writes one ``results_*.xlsx`` per questionnaire found.

:func:`discover_questionnaire_groups`
    Groups ``.xlsx`` files in a folder by the ``config_hash`` stored in
    their hidden ``_meta`` sheet.

:func:`resolve_config`
    Reconstructs a :class:`~umfrage.models.Questionnaire` for a given
    ``config_hash`` by scanning ``*_metadata.yaml`` files or using a
    caller-supplied override.

:func:`collect_group`
    Validates and aggregates a single group of response files into a result
    workbook.

Result workbook layout
----------------------
Row 1   : Title
Row 2   : Collection date and organizer info
Row 3   : Spacer
Row 4   : Column headers — Section | Q-ID | Question | Scale/Comment |
          [Institution A] | [Institution B] | …
Row 5+  : Section sub-headers + one data row per question;
          institution answers fill the remaining columns.
          Cells with missing/required answers are highlighted in the
          configured warning colour.
"""

from __future__ import annotations

import datetime
import json
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from umfrage.config_loader import ConfigError
from umfrage.models import AnswerType, Questionnaire, StyleConfig
from umfrage.styles import (
    apply_result_header_style,
    apply_section_style,
    make_fill,
    make_font,
    make_thin_border,
)
from umfrage.translator import Translator
from umfrage.validator import META_SHEET, ValidationResult, validate_response

# Result files begin with this prefix so the collector can skip them on re-runs.
RESULTS_PREFIX = "results_"


@dataclass
class CollectionSummary:
    """Summary of one questionnaire group's collection run."""

    questionnaire_id: str
    questionnaire_title: str
    total_files: int
    valid_count: int
    skipped_count: int
    skipped_files: list[tuple[Path, list[str]]] = field(default_factory=list)
    output_path: Path | None = None


def collect_all(
    folder: Path,
    style: StyleConfig,
    output_dir: Path,
    config_override: Questionnaire | None = None,
) -> list[CollectionSummary]:
    """Process all response files in *folder*, grouped by questionnaire identity.

    For each group of files sharing the same ``config_hash``:

    1. Resolve the questionnaire config (from ``*_metadata.yaml`` or
       *config_override*).
    2. Validate every file in the group.
    3. Aggregate valid responses into ``results_{qid}_{date}.xlsx`` in
       *output_dir*.

    Args:
        folder: Directory containing returned ``.xlsx`` files (and
            optionally ``*_metadata.yaml`` files).
        style: Excel appearance and protection configuration.
        output_dir: Directory where result files are written.
        config_override: If supplied, this questionnaire is used for **all**
            groups instead of auto-discovering from metadata files.

    Returns:
        A :class:`CollectionSummary` per questionnaire group found.
    """
    groups = discover_questionnaire_groups(folder)
    summaries: list[CollectionSummary] = []

    for config_hash, response_paths in groups.items():
        try:
            questionnaire = resolve_config(config_hash, folder, config_override)
        except ConfigError as exc:
            print(
                f"[WARNING] Cannot resolve config for hash {config_hash[:12]}…: {exc}\n"
                f"          Skipping {len(response_paths)} file(s)."
            )
            continue

        qid = questionnaire.questionnaire_id()
        date_str = datetime.date.today().isoformat()
        result_filename = f"{RESULTS_PREFIX}{qid}_{date_str}.xlsx"
        result_path = output_dir / result_filename
        output_dir.mkdir(parents=True, exist_ok=True)

        summary = collect_group(response_paths, questionnaire, style, result_path)
        summaries.append(summary)

    return summaries


def discover_questionnaire_groups(folder: Path) -> dict[str, list[Path]]:
    """Scan *folder* for response ``.xlsx`` files and group by ``config_hash``.

    Files whose names start with ``results_`` are skipped (they are output
    files from a previous collection run). Files that cannot be opened or
    lack a valid ``_meta`` sheet are skipped with a printed warning.

    Returns:
        ``{config_hash: [path, …]}`` mapping.
    """
    groups: dict[str, list[Path]] = {}

    for xlsx_path in sorted(folder.glob("*.xlsx")):
        if xlsx_path.name.startswith(RESULTS_PREFIX):
            continue

        config_hash = _read_config_hash(xlsx_path)
        if config_hash is None:
            print(
                f"[WARNING] Skipping '{xlsx_path.name}': "
                "cannot read config hash from '_meta' sheet."
            )
            continue

        groups.setdefault(config_hash, []).append(xlsx_path)

    return groups


def resolve_config(
    config_hash: str,
    folder: Path,
    config_override: Questionnaire | None = None,
) -> Questionnaire:
    """Reconstruct the questionnaire config for *config_hash*.

    Resolution order:

    1. *config_override* — used as-is (hash is not checked against it so
       that manually supplied configs always win).
    2. ``*_metadata.yaml`` files in *folder* whose embedded ``config_hash``
       matches.

    Raises:
        :exc:`~umfrage.config_loader.ConfigError`: if no config can be found.
    """
    if config_override is not None:
        return config_override

    for metadata_path in sorted(folder.glob("*_metadata.yaml")):
        try:
            data = yaml.safe_load(metadata_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        if not isinstance(data, dict):
            continue

        if data.get("config_hash") == config_hash:
            questionnaire_data = data.get("questionnaire")
            if questionnaire_data:
                from umfrage.config_loader import questionnaire_from_dict
                return questionnaire_from_dict(questionnaire_data)

    raise ConfigError(
        f"No questionnaire config found for hash {config_hash[:12]}…  "
        "Either run 'umfrage generate --metadata-file' to create a "
        "*_metadata.yaml companion file, or pass --config."
    )


def collect_group(
    response_paths: list[Path],
    questionnaire: Questionnaire,
    style: StyleConfig,
    output_path: Path,
) -> CollectionSummary:
    """Validate and aggregate one group of response files into a result workbook.

    Args:
        response_paths: Paths to response ``.xlsx`` files for this group.
        questionnaire: The questionnaire config for this group.
        style: Excel appearance configuration.
        output_path: Destination for the aggregated result ``.xlsx``.

    Returns:
        A :class:`CollectionSummary` for this group.
    """
    summary = CollectionSummary(
        questionnaire_id=questionnaire.questionnaire_id(),
        questionnaire_title=questionnaire.title,
        total_files=len(response_paths),
        valid_count=0,
        skipped_count=0,
        output_path=output_path,
    )

    valid_results: list[ValidationResult] = []

    for path in response_paths:
        vr = validate_response(path, questionnaire)
        if vr.is_valid:
            valid_results.append(vr)
            summary.valid_count += 1
        else:
            summary.skipped_count += 1
            summary.skipped_files.append((path, vr.errors))
            print(
                f"[WARNING] Skipping '{path.name}': "
                + "; ".join(vr.errors)
            )

    if valid_results:
        _build_result_workbook(valid_results, questionnaire, style, output_path)

    return summary


# ── Internal helpers ──────────────────────────────────────────────────────────

def _read_config_hash(xlsx_path: Path) -> str | None:
    """Safely read the ``config_hash`` value from a response file's ``_meta`` sheet."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if META_SHEET not in wb.sheetnames:
            return None
        ws_meta = wb[META_SHEET]
        for row in ws_meta.iter_rows(max_col=2, values_only=True):
            if row[0] == "config_hash":
                return str(row[1]) if row[1] else None
        return None
    except Exception:
        return None


def _find_institution_field(questionnaire: Questionnaire) -> str:
    """Return the label of the most likely 'institution' respondent field.

    Prefers a field whose label contains 'institution' or 'organization'.
    Falls back to the first field if none match.
    """
    for rf in questionnaire.respondent_fields:
        if any(kw in rf.label.lower() for kw in ("institution", "organization", "org")):
            return rf.label
    return questionnaire.respondent_fields[0].label if questionnaire.respondent_fields else "Respondent"


def _build_result_workbook(
    valid_results: list[ValidationResult],
    questionnaire: Questionnaire,
    style: StyleConfig,
    output_path: Path,
) -> None:
    """Build and save the aggregated result workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    translator = Translator(questionnaire.language)
    institution_field = _find_institution_field(questionnaire)
    institution_names = [
        vr.respondent_info.get(institution_field) or f"Respondent {i + 1}"
        for i, vr in enumerate(valid_results)
    ]

    # Fixed columns: Section | Q-ID | Question | Scale/Comment | [respondents…]
    fixed_cols = 4
    total_cols = fixed_cols + len(valid_results)
    warning_fill = PatternFill(fill_type="solid", fgColor=style.warning_color)

    row = 1

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 28
    title_cell = ws.cell(
        row=row, column=1,
        value=f"{questionnaire.title} {translator.t('result_title_suffix')}",
    )
    apply_result_header_style(title_cell, style)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    row += 1

    # ── Row 2: Collection metadata ────────────────────────────────────────────
    ws.row_dimensions[row].height = 16
    org = questionnaire.organizer
    meta_text = (
        f"{translator.t('result_collected')}: {datetime.date.today().isoformat()}  |  "
        f"{translator.t('result_organizer')}: {org.name}, {org.institution}"
    )
    meta_cell = ws.cell(row=row, column=1, value=meta_text)
    meta_cell.font = make_font(style.result_header, size_override=9)
    meta_cell.fill = make_fill(style.result_header.background_color)
    meta_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    row += 1

    # ── Row 3: Spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 8
    row += 1

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[row].height = 40
    col_headers = [
        translator.t("result_col_section"),
        translator.t("result_col_qid"),
        translator.t("result_col_question"),
        translator.t("result_col_scale_comment"),
    ] + institution_names
    col_widths = [22, 10, 52, 30] + [22] * len(valid_results)

    for col_idx, (header, width) in enumerate(zip(col_headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_result_header_style(cell, style)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    row += 1

    # ── Row 5: Source-file row ────────────────────────────────────────────────
    # Shows the filename of each response file for quality control, with a
    # clickable file:// hyperlink back to the original .xlsx on disk.
    _SOURCE_BG = "F0F4F8"   # very light blue-grey
    _LINK_COLOR = "0563C1"  # standard hyperlink blue

    ws.row_dimensions[row].height = 16
    label_cell = ws.cell(
        row=row, column=1,
        value=translator.t("result_col_source_file"),
    )
    label_cell.font = Font(size=8, italic=True, color="666666")
    label_cell.fill = make_fill(_SOURCE_BG)
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    label_cell.border = make_thin_border()

    for c in range(2, fixed_cols + 1):
        bg_cell = ws.cell(row=row, column=c)
        bg_cell.fill = make_fill(_SOURCE_BG)
        bg_cell.border = make_thin_border()

    for resp_idx, vr in enumerate(valid_results):
        col_idx = fixed_cols + 1 + resp_idx
        fname_cell = ws.cell(row=row, column=col_idx, value=vr.path.name)
        fname_cell.font = Font(size=8, color=_LINK_COLOR, underline="single")
        fname_cell.fill = make_fill(_SOURCE_BG)
        fname_cell.alignment = Alignment(horizontal="left", vertical="center")
        fname_cell.border = make_thin_border()
        try:
            # Use a path relative to the results file so the link works on any
            # machine that has the folder structure intact, without exposing
            # absolute filesystem paths.
            rel = os.path.relpath(
                vr.path.resolve(), output_path.parent.resolve()
            ).replace(os.sep, "/")
            fname_cell.hyperlink = rel
        except (ValueError, OSError):
            pass  # Path cannot be resolved on this system; skip hyperlink
    row += 1

    # ── Question rows ─────────────────────────────────────────────────────────
    question_idx = 0

    for section in questionnaire.sections:
        # Section sub-header
        ws.row_dimensions[row].height = 16
        section_cell = ws.cell(row=row, column=1, value=f"  {section.title}")
        apply_section_style(section_cell, style)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        row += 1

        for q in section.questions:
            alternate = question_idx % 2 == 1
            q_color = (
                style.question_row.alternate_color
                if (alternate and style.question_row.alternate_color)
                else style.question_row.background_color
            )
            q_fill = make_fill(q_color)
            q_font = make_font(style.question_row)
            border = make_thin_border()

            ws.row_dimensions[row].height = 32

            # Comment/scale hint text
            comment_text = q.comment or ""
            if not comment_text and q.answer.type == AnswerType.SCALE:
                if q.answer.min_value is not None:
                    comment_text = f"{q.answer.min_value}–{q.answer.max_value}"

            for col_idx, value in enumerate(
                [section.title, q.id, q.text, comment_text], start=1
            ):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.fill = q_fill
                cell.font = q_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

            # Respondent answer columns
            for resp_idx, vr in enumerate(valid_results):
                col_idx = fixed_cols + 1 + resp_idx
                answer_value = vr.answers.get(q.id)
                cell = ws.cell(row=row, column=col_idx, value=answer_value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                # Highlight missing required answers in warning colour
                is_empty = answer_value is None or str(answer_value).strip() == ""
                cell.fill = warning_fill if (is_empty and q.required) else make_fill("FFFFFF")

            question_idx += 1
            row += 1

    wb.save(output_path)
