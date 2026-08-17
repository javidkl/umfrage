"""Tests for umfrage.cli — focused on the 'list' command and '--survey' flag."""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest
import yaml
from click.testing import CliRunner

from umfrage.cli import main
from umfrage.generator import generate_questionnaire, write_metadata_file
from umfrage.models import Questionnaire


# ── helpers / fixtures ────────────────────────────────────────────────────────

def _fill_response(xlsx_path: Path, name: str, institution: str, answers: dict) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb["Questionnaire"]
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        cell_text = str(val)
        if cell_text == "Name:":
            ws.cell(row=row, column=2).value = name
        elif cell_text == "Institution:":
            ws.cell(row=row, column=2).value = institution
        elif cell_text in answers:
            ws.cell(row=row, column=3).value = answers[cell_text]
    wb.save(xlsx_path)


SAMPLE_ANSWERS = {"G.Q1": 4, "G.Q2": "Yes", "G.Q3": "Great!", "G.Q4": "Good", "T.Q1": 8}


@pytest.fixture()
def config_yaml_path(tmp_path: Path, sample_questionnaire: Questionnaire) -> Path:
    """Write sample_questionnaire to a temp YAML file and return the path."""
    yaml_path = tmp_path / "questionnaire.yaml"
    yaml_path.write_text(
        yaml.dump(sample_questionnaire.model_dump(mode="json"), allow_unicode=True),
        encoding="utf-8",
    )
    return yaml_path


# ── helpers ───────────────────────────────────────────────────────────────────

def _fill_response(xlsx_path: Path, name: str, institution: str, answers: dict) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb["Questionnaire"]
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        cell_text = str(val)
        if cell_text == "Name:":
            ws.cell(row=row, column=2).value = name
        elif cell_text == "Institution:":
            ws.cell(row=row, column=2).value = institution
        elif cell_text in answers:
            ws.cell(row=row, column=3).value = answers[cell_text]
    wb.save(xlsx_path)


SAMPLE_ANSWERS = {"G.Q1": 4, "G.Q2": "Yes", "G.Q3": "Great!", "G.Q4": "Good", "T.Q1": 8}


# ── umfrage list ──────────────────────────────────────────────────────────────

class TestCmdList:
    def test_shows_group_info(
        self, tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
    ) -> None:
        folder = tmp_path / "responses"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base, config_file="my-survey.yaml")
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = folder / "resp.xlsx"
        shutil.copy(base, resp)
        _fill_response(resp, "Alice", "Uni A", SAMPLE_ANSWERS)

        runner = CliRunner()
        result = runner.invoke(main, ["list", str(folder)])
        assert result.exit_code == 0
        assert sample_questionnaire.questionnaire_id() in result.output
        assert sample_questionnaire.title in result.output
        assert sample_questionnaire.config_hash()[:12] in result.output
        assert "my-survey.yaml" in result.output

    def test_empty_folder_message(self, tmp_path: Path) -> None:
        folder = tmp_path / "empty"
        folder.mkdir()
        runner = CliRunner()
        result = runner.invoke(main, ["list", str(folder)])
        assert result.exit_code == 0
        assert "No questionnaire response files found" in result.output

    def test_two_groups_shows_two_rows(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = tmp_path / "mixed"
        folder.mkdir()
        for q_obj, answers in [
            (sample_questionnaire, SAMPLE_ANSWERS),
            (other_questionnaire, {"F.Q1": 2}),
        ]:
            base = tmp_path / f"base_{q_obj.questionnaire_id()}.xlsx"
            generate_questionnaire(q_obj, sample_style, base)
            write_metadata_file(q_obj, folder / f"{q_obj.questionnaire_id()}_metadata.yaml")
            resp = folder / f"resp_{q_obj.questionnaire_id()}.xlsx"
            shutil.copy(base, resp)
            _fill_response(resp, "Tester", "Org", answers)

        runner = CliRunner()
        result = runner.invoke(main, ["list", str(folder)])
        assert result.exit_code == 0
        assert "2 questionnaire group(s)" in result.output
        assert sample_questionnaire.questionnaire_id() in result.output
        assert other_questionnaire.questionnaire_id() in result.output

    def test_unresolvable_group_flagged(
        self, tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
    ) -> None:
        folder = tmp_path / "no_meta"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        resp = folder / "resp.xlsx"
        shutil.copy(base, resp)

        runner = CliRunner()
        result = runner.invoke(main, ["list", str(folder)])
        assert result.exit_code == 0
        assert "no metadata" in result.output

    def test_files_flag_lists_filenames(
        self, tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
    ) -> None:
        folder = tmp_path / "responses"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        for name in ("alice.xlsx", "bob.xlsx"):
            resp = folder / name
            shutil.copy(base, resp)
            _fill_response(resp, name.split(".")[0].title(), "Org", SAMPLE_ANSWERS)

        runner = CliRunner()
        result = runner.invoke(main, ["list", "--files", str(folder)])
        assert result.exit_code == 0
        assert "alice.xlsx" in result.output
        assert "bob.xlsx" in result.output

    def test_no_files_flag_omits_filenames(
        self, tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
    ) -> None:
        folder = tmp_path / "responses"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = folder / "alice.xlsx"
        shutil.copy(base, resp)

        runner = CliRunner()
        result = runner.invoke(main, ["list", str(folder)])
        assert result.exit_code == 0
        assert "alice.xlsx" not in result.output
        assert "--files" in result.output  # tip is shown


# ── umfrage collect --survey ──────────────────────────────────────────────────

class TestCmdCollectSurveyFlag:
    def _mixed_folder(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> Path:
        folder = tmp_path / "mixed"
        folder.mkdir()
        for q_obj, answers in [
            (sample_questionnaire, SAMPLE_ANSWERS),
            (other_questionnaire, {"F.Q1": 2}),
        ]:
            base = tmp_path / f"base_{q_obj.questionnaire_id()}.xlsx"
            generate_questionnaire(q_obj, sample_style, base)
            write_metadata_file(q_obj, folder / f"{q_obj.questionnaire_id()}_metadata.yaml")
            resp = folder / f"resp_{q_obj.questionnaire_id()}.xlsx"
            shutil.copy(base, resp)
            _fill_response(resp, "Tester", "Org", answers)
        return folder

    def test_survey_filter_produces_one_result(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = self._mixed_folder(tmp_path, sample_questionnaire, other_questionnaire, sample_style)
        out_dir = tmp_path / "out"
        runner = CliRunner()
        result = runner.invoke(
            main,
            [
                "collect", str(folder),
                "--output-dir", str(out_dir),
                "--skip-invalid",
                "--survey", sample_questionnaire.questionnaire_id(),
            ],
        )
        assert result.exit_code == 0
        result_files = list(out_dir.glob("results_*.xlsx"))
        assert len(result_files) == 1
        assert sample_questionnaire.questionnaire_id() in result_files[0].name

    def test_survey_ignored_when_config_given(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
        config_yaml_path: Path,
    ) -> None:
        folder = tmp_path / "responses"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = folder / "resp.xlsx"
        shutil.copy(base, resp)
        _fill_response(resp, "Alice", "Uni A", SAMPLE_ANSWERS)

        out_dir = tmp_path / "out"
        runner = CliRunner()
        result = runner.invoke(
            main,
            [
                "collect", str(folder),
                "--config", str(config_yaml_path),
                "--output-dir", str(out_dir),
                "--skip-invalid",
                "--survey", "whatever",
            ],
        )
        assert result.exit_code == 0
        assert "--survey is ignored" in result.output

    def test_slug_collision_exits_with_error(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = tmp_path / "collision"
        folder.mkdir()
        twin = sample_questionnaire.model_copy(update={"version": "2.0"})

        for q_obj in (sample_questionnaire, twin):
            base = tmp_path / f"base_{q_obj.config_hash()[:8]}.xlsx"
            generate_questionnaire(q_obj, sample_style, base)
            write_metadata_file(q_obj, folder / f"{q_obj.config_hash()[:8]}_metadata.yaml")
            resp = folder / f"resp_{q_obj.config_hash()[:8]}.xlsx"
            shutil.copy(base, resp)
            _fill_response(resp, "User", "Org", SAMPLE_ANSWERS)

        out_dir = tmp_path / "out"
        runner = CliRunner()
        result = runner.invoke(
            main,
            [
                "collect", str(folder),
                "--output-dir", str(out_dir),
                "--skip-invalid",
                "--survey", sample_questionnaire.questionnaire_id(),
            ],
        )
        assert result.exit_code == 1
        assert "ambiguous" in result.output
