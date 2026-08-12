"""Tests for umfrage.collector."""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from umfrage.collector import (
    CollectionSummary,
    collect_all,
    collect_group,
    discover_questionnaire_groups,
    resolve_config,
)
from umfrage.config_loader import ConfigError
from umfrage.generator import generate_questionnaire, write_metadata_file
from umfrage.models import (
    AnswerConfig,
    AnswerType,
    OrganizerInfo,
    Question,
    Questionnaire,
    RespondentField,
    Section,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill_response(
    xlsx_path: Path, name: str, institution: str, answers: dict
) -> None:
    """Write respondent info and answers into a copy of the base questionnaire."""
    wb = load_workbook(xlsx_path)
    ws = wb["Questionnaire"]
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        cell_text = str(val)
        if cell_text == "Name:":
            ws.cell(row=row, column=2).value = name
        elif cell_text == "Institution:":
            ws.cell(row=row, column=2).value = institution
        elif cell_text in answers:
            ws.cell(row=row, column=3).value = answers[cell_text]
    wb.save(xlsx_path)


SAMPLE_ANSWERS = {"G.Q1": 4, "G.Q2": "Yes", "G.Q3": "Great work!", "T.Q1": 8}


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture()
def responses_folder(
    tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
) -> Path:
    """Folder with 2 valid filled response files + metadata yaml."""
    folder = tmp_path / "responses"
    folder.mkdir()

    base_xlsx = tmp_path / "base.xlsx"
    generate_questionnaire(sample_questionnaire, sample_style, base_xlsx)
    write_metadata_file(
        sample_questionnaire,
        folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
    )

    for i, (name, inst) in enumerate([("Alice", "Univ A"), ("Bob", "Univ B")], start=1):
        resp_path = folder / f"response_{i}.xlsx"
        shutil.copy(base_xlsx, resp_path)
        _fill_response(resp_path, name, inst, SAMPLE_ANSWERS)

    return folder


@pytest.fixture()
def other_questionnaire() -> Questionnaire:
    """A second questionnaire with a different config hash."""
    return Questionnaire(
        title="Other Survey",
        organizer=OrganizerInfo(name="Bob", institution="Org2", email="b@org2.com"),
        respondent_fields=[
            RespondentField(label="Name"),
            RespondentField(label="Institution"),
        ],
        sections=[
            Section(
                title="Feedback",
                questions=[
                    Question(
                        id="F.Q1",
                        text="Overall rating",
                        answer=AnswerConfig(
                            type=AnswerType.SCALE, min_value=1, max_value=3
                        ),
                    )
                ],
            )
        ],
    )


# ── discover_questionnaire_groups ─────────────────────────────────────────────

class TestDiscoverGroups:
    def test_two_responses_form_one_group(self, responses_folder: Path) -> None:
        groups = discover_questionnaire_groups(responses_folder)
        assert len(groups) == 1
        paths = list(groups.values())[0]
        assert len(paths) == 2

    def test_result_files_are_skipped(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        # Place a results file in the folder; it must not be picked up
        results_file = responses_folder / "results_test_2024-01-01.xlsx"
        shutil.copy(list(responses_folder.glob("response_*.xlsx"))[0], results_file)
        groups = discover_questionnaire_groups(responses_folder)
        paths = list(groups.values())[0]
        assert results_file not in paths

    def test_corrupted_file_is_skipped(self, responses_folder: Path) -> None:
        bad = responses_folder / "bad.xlsx"
        bad.write_bytes(b"not an xlsx")
        groups = discover_questionnaire_groups(responses_folder)
        # Corrupted file should not appear in any group
        all_paths = [p for paths in groups.values() for p in paths]
        assert bad not in all_paths

    def test_empty_folder_returns_empty(self, tmp_path: Path) -> None:
        folder = tmp_path / "empty"
        folder.mkdir()
        assert discover_questionnaire_groups(folder) == {}


# ── resolve_config ────────────────────────────────────────────────────────────

class TestResolveConfig:
    def test_resolves_from_metadata_yaml(
        self, responses_folder: Path, sample_questionnaire: Questionnaire
    ) -> None:
        q = resolve_config(sample_questionnaire.config_hash(), responses_folder)
        assert q.title == sample_questionnaire.title
        assert q.config_hash() == sample_questionnaire.config_hash()

    def test_config_override_takes_precedence(
        self, responses_folder: Path, sample_questionnaire: Questionnaire
    ) -> None:
        q = resolve_config("any_hash", responses_folder, config_override=sample_questionnaire)
        assert q.title == sample_questionnaire.title

    def test_unknown_hash_raises_config_error(self, responses_folder: Path) -> None:
        with pytest.raises(ConfigError, match="No questionnaire config"):
            resolve_config("deadbeef" * 8, responses_folder)

    def test_empty_folder_raises_config_error(self, tmp_path: Path) -> None:
        folder = tmp_path / "empty"
        folder.mkdir()
        with pytest.raises(ConfigError):
            resolve_config("any_hash", folder)


# ── collect_all ───────────────────────────────────────────────────────────────

class TestCollectAll:
    def test_produces_one_summary(
        self, responses_folder: Path, sample_style
    ) -> None:
        out_dir = responses_folder / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        assert len(summaries) == 1

    def test_summary_counts_correct(
        self, responses_folder: Path, sample_style
    ) -> None:
        out_dir = responses_folder / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        s = summaries[0]
        assert s.valid_count == 2
        assert s.skipped_count == 0
        assert s.total_files == 2

    def test_result_file_created(
        self, responses_folder: Path, sample_style
    ) -> None:
        out_dir = responses_folder / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        assert summaries[0].output_path is not None
        assert summaries[0].output_path.exists()

    def test_invalid_file_skipped_during_discovery(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        # A corrupted file is dropped by discover_questionnaire_groups (can't
        # read _meta), so it never enters any group. The two valid files should
        # still produce a successful result.
        bad = responses_folder / "bad.xlsx"
        bad.write_bytes(b"not xlsx")
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        # The bad file was silently dropped at discovery; valid files succeed.
        assert len(summaries) == 1
        assert summaries[0].valid_count == 2
        assert summaries[0].skipped_count == 0

    def test_two_questionnaires_produce_two_results(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = tmp_path / "mixed"
        folder.mkdir()

        for q_obj, answers in [
            (sample_questionnaire, SAMPLE_ANSWERS),
            (other_questionnaire, {"F.Q1": 2}),
        ]:
            base = tmp_path / f"base_{q_obj.questionnaire_id()}.xlsx"
            generate_questionnaire(q_obj, sample_style, base)
            write_metadata_file(
                q_obj, folder / f"{q_obj.questionnaire_id()}_metadata.yaml"
            )
            resp = folder / f"resp_{q_obj.questionnaire_id()}.xlsx"
            shutil.copy(base, resp)
            _fill_response(resp, "Tester", "Org", answers)

        out_dir = tmp_path / "results"
        summaries = collect_all(folder, sample_style, out_dir)
        assert len(summaries) == 2

        result_files = sorted(out_dir.glob("results_*.xlsx"))
        assert len(result_files) == 2

    def test_output_dir_created_if_missing(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "new" / "dir"
        collect_all(responses_folder, sample_style, out_dir)
        assert out_dir.exists()


# ── Result workbook content ───────────────────────────────────────────────────

class TestResultWorkbook:
    def _run_and_open(self, responses_folder, sample_style, tmp_path):
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path, data_only=True)
        return wb["Results"]

    def test_results_sheet_exists(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path)
        assert "Results" in wb.sheetnames

    def test_institution_names_in_header_row(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        ws = self._run_and_open(responses_folder, sample_style, tmp_path)
        # Row 4 is the column header row (fixed cols: Section, Q-ID, Question, Scale/Comment)
        header_values = [ws.cell(row=4, column=c).value for c in range(1, 8)]
        assert "Univ A" in header_values or "Alice" in header_values
        assert "Univ B" in header_values or "Bob" in header_values

    def test_question_ids_in_result(
        self, responses_folder: Path, sample_style, tmp_path: Path,
        sample_questionnaire: Questionnaire
    ) -> None:
        ws = self._run_and_open(responses_folder, sample_style, tmp_path)
        col_b = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        for q in sample_questionnaire.all_questions():
            assert q.id in col_b, f"Question ID '{q.id}' not found in result column B"

    def test_title_in_row_one(
        self, responses_folder: Path, sample_style, tmp_path: Path,
        sample_questionnaire: Questionnaire
    ) -> None:
        ws = self._run_and_open(responses_folder, sample_style, tmp_path)
        title_cell = ws.cell(row=1, column=1).value
        assert sample_questionnaire.title in str(title_cell)


class TestSourceFileRow:
    """Row 5 of the result sheet must list response filenames with hyperlinks."""

    def _result_ws(self, responses_folder, sample_style, tmp_path):
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path, data_only=True)
        return wb["Results"]

    def test_source_file_row_contains_filenames(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        ws = self._result_ws(responses_folder, sample_style, tmp_path)
        # Row 5 is the source-file row (row 4 = institution names)
        row5 = [ws.cell(row=5, column=c).value for c in range(1, 8)]
        assert any(
            str(v).endswith(".xlsx") for v in row5 if v is not None
        ), f"No .xlsx filename found in row 5: {row5}"

    def test_source_file_row_has_both_response_filenames(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        ws = self._result_ws(responses_folder, sample_style, tmp_path)
        row5 = [str(ws.cell(row=5, column=c).value or "") for c in range(1, 8)]
        assert any("response_1" in v for v in row5), "response_1.xlsx not found in row 5"
        assert any("response_2" in v for v in row5), "response_2.xlsx not found in row 5"

    def test_source_file_label_in_col_one(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        ws = self._result_ws(responses_folder, sample_style, tmp_path)
        label = ws.cell(row=5, column=1).value
        # Label must be non-empty (translated "Source file" / "Quelldatei")
        assert label is not None and str(label).strip() != ""

    def test_filename_cells_have_hyperlinks(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        """Filename cells in the source-file row must carry file:// hyperlinks."""
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        # Reload *without* data_only so hyperlink metadata is available
        wb = load_workbook(summaries[0].output_path)
        ws = wb["Results"]
        # Fixed cols: Section(1), Q-ID(2), Question(3), Scale(4) — inst cols start at 5
        hyperlinks_found = []
        for col in range(5, 5 + 2):  # 2 respondents
            cell = ws.cell(row=5, column=col)
            hyperlinks_found.append(cell.hyperlink is not None)
        assert all(hyperlinks_found), (
            f"Expected hyperlinks on all filename cells, got: {hyperlinks_found}"
        )

    def test_filename_cells_are_left_aligned(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path)
        ws = wb["Results"]
        for col in range(5, 5 + 2):
            cell = ws.cell(row=5, column=col)
            assert cell.alignment.horizontal == "left", (
                f"Expected left alignment on filename cell col {col}, "
                f"got '{cell.alignment.horizontal}'"
            )

    def test_hyperlinks_point_to_xlsx_files(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path)
        ws = wb["Results"]
        for col in range(5, 5 + 2):
            cell = ws.cell(row=5, column=col)
            assert cell.hyperlink is not None
            target = str(cell.hyperlink.target)
            assert target.endswith(".xlsx"), (
                f"Hyperlink target does not end with .xlsx: {target}"
            )
            # Must be relative — no absolute path or file:// URI leaked
            assert not target.startswith("file://"), (
                f"Hyperlink should be relative, not a file:// URI: {target}"
            )
            assert not target.startswith("/"), (
                f"Hyperlink should be relative, not absolute: {target}"
            )

    def test_german_source_file_label_translated(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        """German questionnaire must show 'Quelldatei' as the row label."""
        import shutil
        from umfrage.generator import generate_questionnaire, write_metadata_file

        q = sample_questionnaire.model_copy(update={"language": "de"})
        folder = tmp_path / "de_responses"
        folder.mkdir()
        base = tmp_path / "base_de.xlsx"
        generate_questionnaire(q, sample_style, base)
        write_metadata_file(q, folder / f"{q.questionnaire_id()}_metadata.yaml")

        resp = folder / "antwort_1.xlsx"
        shutil.copy(base, resp)
        # Fill required fields so validation passes
        wb_resp = load_workbook(resp)
        ws_resp = wb_resp["Questionnaire"]
        for row in range(1, ws_resp.max_row + 1):
            v = ws_resp.cell(row=row, column=1).value
            if v == "Name:":
                ws_resp.cell(row=row, column=2).value = "Hans"
            elif v == "Institution:":
                ws_resp.cell(row=row, column=2).value = "Uni Berlin"
            elif v == "G.Q1":
                ws_resp.cell(row=row, column=3).value = 3
            elif v == "G.Q2":
                ws_resp.cell(row=row, column=3).value = "Ja"
            elif v == "T.Q1":
                ws_resp.cell(row=row, column=3).value = 5
        wb_resp.save(resp)

        out_dir = tmp_path / "out_de"
        summaries = collect_all(folder, sample_style, out_dir)
        assert summaries, "No collection summaries returned"
        wb_result = load_workbook(summaries[0].output_path, data_only=True)
        ws_result = wb_result["Results"]
        label = str(ws_result.cell(row=5, column=1).value or "")
        assert "Quelldatei" in label, (
            f"Expected 'Quelldatei' for German label, got: '{label}'"
        )
