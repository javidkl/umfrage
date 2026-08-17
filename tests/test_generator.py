"""Tests for umfrage.generator."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from umfrage.generator import generate_questionnaire, write_metadata_file
from umfrage.models import StyleConfig, Questionnaire


class TestGeneratedFile:
    def test_output_file_exists(self, generated_xlsx: Path) -> None:
        assert generated_xlsx.exists()
        assert generated_xlsx.suffix == ".xlsx"

    def test_output_file_is_valid_xlsx(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        assert wb is not None

    def test_questionnaire_sheet_present(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        assert "Questionnaire" in wb.sheetnames

    def test_meta_sheet_present(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        assert "_meta" in wb.sheetnames

    def test_meta_sheet_is_hidden(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        assert wb["_meta"].sheet_state == "hidden"

    def test_output_dir_created_if_missing(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        nested = tmp_path / "a" / "b" / "c"
        out = nested / "q.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, out)
        assert out.exists()


class TestMetaSheet:
    def _read_meta(self, path: Path) -> dict:
        wb = load_workbook(path, data_only=True)
        ws = wb["_meta"]
        return {
            row[0]: row[1]
            for row in ws.iter_rows(max_col=2, values_only=True)
            if row[0]
        }

    def test_meta_has_questionnaire_id(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        meta = self._read_meta(generated_xlsx)
        assert meta["questionnaire_id"] == sample_questionnaire.questionnaire_id()

    def test_meta_has_config_hash(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        meta = self._read_meta(generated_xlsx)
        assert meta["config_hash"] == sample_questionnaire.config_hash()

    def test_meta_question_ids_match(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        meta = self._read_meta(generated_xlsx)
        stored_ids = json.loads(meta["question_ids"])
        expected_ids = [q.id for q in sample_questionnaire.all_questions()]
        assert set(stored_ids) == set(expected_ids)
        assert stored_ids == expected_ids  # order preserved

    def test_meta_has_embedded_model(self, generated_xlsx: Path) -> None:
        meta = self._read_meta(generated_xlsx)
        assert "questionnaire_json" in meta
        model_data = json.loads(meta["questionnaire_json"])
        assert "title" in model_data
        assert "sections" in model_data

    def test_meta_has_version(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        meta = self._read_meta(generated_xlsx)
        assert meta["version"] == sample_questionnaire.version

    def test_meta_has_project_url(self, generated_xlsx: Path) -> None:
        meta = self._read_meta(generated_xlsx)
        assert meta.get("project_url") == "https://github.com/scinnod/umfrage"

    def test_meta_has_config_file_when_passed(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        out = tmp_path / "q.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, out, config_file="my-survey.yaml")
        meta = self._read_meta(out)
        assert meta.get("config_file") == "my-survey.yaml"

    def test_meta_config_file_absent_when_not_passed(self, generated_xlsx: Path) -> None:
        meta = self._read_meta(generated_xlsx)
        assert "config_file" not in meta

    def test_meta_sheet_is_protected(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        assert wb["_meta"].protection.sheet is True

    def test_meta_sheet_password_matches_questionnaire(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        style = StyleConfig(protection_password="test_secret")
        out = tmp_path / "protected.xlsx"
        generate_questionnaire(sample_questionnaire, style, out)
        wb = load_workbook(out)
        assert wb["_meta"].protection.password == wb["Questionnaire"].protection.password


class TestFooter:
    def test_footer_present_by_default(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx, data_only=True)
        ws = wb["Questionnaire"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert any("umfrage" in str(v) for v in col_a if v), (
            "Expected a footer row containing 'umfrage' on the questionnaire sheet"
        )

    def test_footer_contains_github_url(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx, data_only=True)
        ws = wb["Questionnaire"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert any("github.com/scinnod/umfrage" in str(v) for v in col_a if v), (
            "Expected footer to contain the GitHub project URL"
        )

    def test_footer_hidden_when_show_footer_false(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        style = StyleConfig(show_footer=False)
        out = tmp_path / "no_footer.xlsx"
        generate_questionnaire(sample_questionnaire, style, out)
        wb = load_workbook(out, data_only=True)
        ws = wb["Questionnaire"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert not any("umfrage" in str(v) for v in col_a if v), (
            "Footer should not appear when show_footer=False"
        )

    def test_german_footer_uses_erzeugt(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        q = sample_questionnaire.model_copy(update={"language": "de"})
        out = tmp_path / "de_footer.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out, data_only=True)
        ws = wb["Questionnaire"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert any("Erzeugt" in str(v) for v in col_a if v), (
            "German questionnaire should use 'Erzeugt mit umfrage' in the footer"
        )


class TestSheetProtection:
    def test_worksheet_protection_enabled(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        ws = wb["Questionnaire"]
        assert ws.protection.sheet is True

    def test_answer_cells_are_unlocked(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        ws = wb["Questionnaire"]
        unlocked = [
            ws.cell(row=r, column=3).protection.locked
            for r in range(1, ws.max_row + 1)
        ]
        # At least one answer cell (column C) must be explicitly unlocked
        assert False in unlocked, (
            "Expected at least one unlocked cell in column C (answer column)"
        )

    def test_title_row_is_locked(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        ws = wb["Questionnaire"]
        assert ws.cell(row=1, column=1).protection.locked is True

    def test_password_protection_applied(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        style = StyleConfig(protection_password="test_secret")
        out = tmp_path / "protected.xlsx"
        generate_questionnaire(sample_questionnaire, style, out)
        wb = load_workbook(out)
        ws = wb["Questionnaire"]
        # openpyxl stores a hashed password; check it is set (not None/empty)
        assert ws.protection.sheet is True
        assert ws.protection.password is not None
        assert ws.protection.password != ""


class TestDataValidation:
    def test_scale_cells_have_whole_number_validation(
        self, generated_xlsx: Path
    ) -> None:
        wb = load_workbook(generated_xlsx)
        ws = wb["Questionnaire"]
        dv_types = [dv.type for dv in ws.data_validations.dataValidation]
        assert "whole" in dv_types

    def test_yesno_cells_have_list_validation(self, generated_xlsx: Path) -> None:
        wb = load_workbook(generated_xlsx)
        ws = wb["Questionnaire"]
        dv_types = [dv.type for dv in ws.data_validations.dataValidation]
        assert "list" in dv_types


class TestSheetContent:
    def test_all_question_ids_present_in_column_a(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        wb = load_workbook(generated_xlsx, data_only=True)
        ws = wb["Questionnaire"]
        col_a = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
        for q in sample_questionnaire.all_questions():
            assert q.id in col_a, f"Question ID '{q.id}' not found in column A"

    def test_title_in_row_one(self, generated_xlsx: Path, sample_questionnaire) -> None:
        wb = load_workbook(generated_xlsx, data_only=True)
        ws = wb["Questionnaire"]
        assert ws.cell(row=1, column=1).value == sample_questionnaire.title

    def test_respondent_fields_present(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        wb = load_workbook(generated_xlsx, data_only=True)
        ws = wb["Questionnaire"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        for rf in sample_questionnaire.respondent_fields:
            assert (rf.label + ":") in col_a, (
                f"Respondent field label '{rf.label}:' not found in column A"
            )


class TestWriteMetadataFile:
    def test_metadata_file_created(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        meta_path = tmp_path / "q_metadata.yaml"
        write_metadata_file(sample_questionnaire, meta_path)
        assert meta_path.exists()

    def test_metadata_contains_hash(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        meta_path = tmp_path / "q_metadata.yaml"
        write_metadata_file(sample_questionnaire, meta_path)
        data = yaml.safe_load(meta_path.read_text())
        assert data["config_hash"] == sample_questionnaire.config_hash()

    def test_metadata_contains_questionnaire_id(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        meta_path = tmp_path / "q_metadata.yaml"
        write_metadata_file(sample_questionnaire, meta_path)
        data = yaml.safe_load(meta_path.read_text())
        assert data["questionnaire_id"] == sample_questionnaire.questionnaire_id()

    def test_metadata_embeds_full_model(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        meta_path = tmp_path / "q_metadata.yaml"
        write_metadata_file(sample_questionnaire, meta_path)
        data = yaml.safe_load(meta_path.read_text())
        assert "questionnaire" in data
        assert data["questionnaire"]["title"] == sample_questionnaire.title
        assert "sections" in data["questionnaire"]

    def test_metadata_question_ids_ordered(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        meta_path = tmp_path / "q_metadata.yaml"
        write_metadata_file(sample_questionnaire, meta_path)
        data = yaml.safe_load(meta_path.read_text())
        expected = [q.id for q in sample_questionnaire.all_questions()]
        assert data["question_ids"] == expected

    def test_metadata_creates_parent_dirs(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        meta_path = tmp_path / "nested" / "dir" / "q_metadata.yaml"
        write_metadata_file(sample_questionnaire, meta_path)
        assert meta_path.exists()


class TestI18nGeneration:
    """Verify that language settings propagate correctly into the Excel file."""

    def _make_german(self, sample_questionnaire):
        return sample_questionnaire.model_copy(update={"language": "de"})

    def test_german_yesno_dropdown_uses_ja_nein(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        q = self._make_german(sample_questionnaire)
        out = tmp_path / "de.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out)
        ws = wb["Questionnaire"]
        list_dvs = [
            dv for dv in ws.data_validations.dataValidation if dv.type == "list"
        ]
        formulas = [dv.formula1 for dv in list_dvs]
        assert any("Ja" in f for f in formulas), (
            f"Expected 'Ja' in a list-validation formula, got: {formulas}"
        )
        assert not any("Yes" in f for f in formulas), (
            "English 'Yes' should not appear in a German questionnaire dropdown"
        )

    def test_german_column_headers_in_sheet(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        q = self._make_german(sample_questionnaire)
        out = tmp_path / "de.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out, data_only=True)
        ws = wb["Questionnaire"]
        all_values = {
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, 5)
        }
        assert "Frage" in all_values, "German 'Frage' not found in column values"
        assert "Antwort" in all_values, "German 'Antwort' not found in column values"

    def test_german_respondent_header_translated(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        q = self._make_german(sample_questionnaire)
        out = tmp_path / "de.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out, data_only=True)
        ws = wb["Questionnaire"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert any("ANGABEN" in str(v) for v in col_a if v), (
            "German respondent-information header not found"
        )

    def test_english_questionnaire_uses_yes_no(
        self, generated_xlsx: Path
    ) -> None:
        wb = load_workbook(generated_xlsx)
        ws = wb["Questionnaire"]
        list_dvs = [
            dv for dv in ws.data_validations.dataValidation if dv.type == "list"
        ]
        formulas = [dv.formula1 for dv in list_dvs]
        assert any("Yes" in f for f in formulas), (
            "English 'Yes' not found in list-validation formula"
        )


class TestChoicesDataValidation:
    """Verify that CHOICES questions produce correct Excel data-validation objects."""

    def _make_choices_questionnaire(self, choice_lists=None, questions=None):
        from umfrage.models import OrganizerInfo, RespondentField, Section
        return Questionnaire(
            title="Choices Test",
            choice_lists=choice_lists or {},
            organizer=OrganizerInfo(name="A", institution="B", email="a@b.org"),
            respondent_fields=[RespondentField(label="Name")],
            sections=[Section(title="S1", questions=questions)],
        )

    def test_choices_cell_has_list_validation(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        out = tmp_path / "choices.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, out)
        wb = load_workbook(out)
        ws = wb["Questionnaire"]
        list_dvs = [dv for dv in ws.data_validations.dataValidation if dv.type == "list"]
        # At least two list DVs: one for yes_no, one for choices
        assert len(list_dvs) >= 2

    def test_choices_dv_formula_is_quoted_csv(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        out = tmp_path / "choices.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, out)
        wb = load_workbook(out)
        ws = wb["Questionnaire"]
        list_dvs = [dv for dv in ws.data_validations.dataValidation if dv.type == "list"]
        # The choices DV formula must be a double-quoted comma-separated string
        choices_formulas = [f for f in (dv.formula1 for dv in list_dvs) if "Poor" in f]
        assert choices_formulas, "Expected a list-DV formula containing 'Poor'"
        formula = choices_formulas[0]
        assert formula.startswith('"') and formula.endswith('"')
        assert "," in formula

    def test_identical_choice_lists_share_one_dv_object(
        self, tmp_path: Path, sample_style
    ) -> None:
        from umfrage.models import AnswerConfig, AnswerType, Question
        opts = ["Alpha", "Beta", "Gamma"]
        q = self._make_choices_questionnaire(
            questions=[
                Question(id="Q1", text="Q1", answer=AnswerConfig(type=AnswerType.CHOICES, choices=opts)),
                Question(id="Q2", text="Q2", answer=AnswerConfig(type=AnswerType.CHOICES, choices=opts)),
            ]
        )
        out = tmp_path / "shared.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out)
        ws = wb["Questionnaire"]
        choices_dvs = [
            dv for dv in ws.data_validations.dataValidation
            if dv.type == "list" and "Alpha" in dv.formula1
        ]
        assert len(choices_dvs) == 1, (
            "Two questions with identical choices should share one DV object"
        )

    def test_different_choice_lists_get_separate_dv_objects(
        self, tmp_path: Path, sample_style
    ) -> None:
        from umfrage.models import AnswerConfig, AnswerType, Question
        q = self._make_choices_questionnaire(
            questions=[
                Question(id="Q1", text="Q1", answer=AnswerConfig(
                    type=AnswerType.CHOICES, choices=["A", "B"])),
                Question(id="Q2", text="Q2", answer=AnswerConfig(
                    type=AnswerType.CHOICES, choices=["X", "Y", "Z"])),
            ]
        )
        out = tmp_path / "different.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out)
        ws = wb["Questionnaire"]
        choices_dvs = [
            dv for dv in ws.data_validations.dataValidation if dv.type == "list"
        ]
        formulas = [dv.formula1 for dv in choices_dvs]
        assert any("A" in f and "B" in f for f in formulas)
        assert any("X" in f and "Y" in f for f in formulas)

    def test_choices_comment_shows_options_by_default(
        self, tmp_path: Path, sample_style
    ) -> None:
        from umfrage.models import AnswerConfig, AnswerType, Question
        q = self._make_choices_questionnaire(
            questions=[Question(
                id="Q1", text="Q1",
                answer=AnswerConfig(type=AnswerType.CHOICES, choices=["Red", "Green", "Blue"]),
            )]
        )
        out = tmp_path / "comment.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out, data_only=True)
        ws = wb["Questionnaire"]
        comments = [
            ws.cell(row=r, column=4).value
            for r in range(1, ws.max_row + 1)
        ]
        assert any(
            v and "Red" in str(v) for v in comments
        ), "Choice options should appear in the comment column by default"

    def test_choices_comment_suppressed_when_opted_out(
        self, tmp_path: Path, sample_style
    ) -> None:
        from umfrage.models import AnswerConfig, AnswerType, Question
        q = self._make_choices_questionnaire(
            questions=[Question(
                id="Q1", text="Q1",
                answer=AnswerConfig(
                    type=AnswerType.CHOICES,
                    choices=["Red", "Green", "Blue"],
                    show_choices_in_comment=False,
                    description="Pick a colour",
                ),
            )]
        )
        out = tmp_path / "no_comment.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out, data_only=True)
        ws = wb["Questionnaire"]
        comments = [
            ws.cell(row=r, column=4).value
            for r in range(1, ws.max_row + 1)
        ]
        assert not any(v and "Red" in str(v) for v in comments), (
            "Choice options should not appear in comment column when show_choices_in_comment=False"
        )
        assert any(v and "Pick a colour" in str(v) for v in comments), (
            "description should still appear in comment column when show_choices_in_comment=False"
        )

    def test_explicit_comment_overrides_choices_list(
        self, tmp_path: Path, sample_style
    ) -> None:
        """comment is a complete override for CHOICES, same as for scale/yes_no/freetext."""
        from umfrage.models import AnswerConfig, AnswerType, Question
        q = self._make_choices_questionnaire(
            questions=[Question(
                id="Q1", text="Q1",
                comment="Custom scenario text",
                answer=AnswerConfig(
                    type=AnswerType.CHOICES,
                    choices=["Red", "Green", "Blue"],
                    show_choices_in_comment=True,  # has no effect when comment is set
                ),
            )]
        )
        out = tmp_path / "comment_override.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out, data_only=True)
        ws = wb["Questionnaire"]
        comments = [ws.cell(row=r, column=4).value for r in range(1, ws.max_row + 1)]
        assert any(v and "Custom scenario text" in str(v) for v in comments), (
            "explicit comment should appear"
        )
        assert not any(v and "Red" in str(v) for v in comments), (
            "choices list must not be appended when comment is set (complete override)"
        )

    def test_comment_override_takes_full_priority(
        self, tmp_path: Path, sample_style
    ) -> None:
        """comment overrides choices list regardless of show_choices_in_comment=False."""
        from umfrage.models import AnswerConfig, AnswerType, Question
        q = self._make_choices_questionnaire(
            questions=[Question(
                id="Q1", text="Q1",
                comment="Scenario description",
                answer=AnswerConfig(
                    type=AnswerType.CHOICES,
                    choices=["Red", "Green", "Blue"],
                    show_choices_in_comment=False,
                ),
            )]
        )
        out = tmp_path / "comment_no_list.xlsx"
        generate_questionnaire(q, sample_style, out)
        wb = load_workbook(out, data_only=True)
        ws = wb["Questionnaire"]
        comments = [ws.cell(row=r, column=4).value for r in range(1, ws.max_row + 1)]
        assert any(v and "Scenario description" in str(v) for v in comments), (
            "explicit comment should appear"
        )
        assert not any(v and "Red" in str(v) for v in comments), (
            "choices list must not appear when comment is set"
        )
