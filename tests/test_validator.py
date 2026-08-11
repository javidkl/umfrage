"""Tests for umfrage.validator."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from umfrage.validator import validate_response


# ── Helper ────────────────────────────────────────────────────────────────────

def _fill_xlsx(path: Path, respondent: dict[str, str], answers: dict[str, object]) -> None:
    """Write respondent info and question answers into a generated questionnaire xlsx."""
    wb = load_workbook(path)
    ws = wb["Questionnaire"]
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        cell_text = str(val)
        # Respondent field rows: "Label:"  value in col B
        for label, field_value in respondent.items():
            if cell_text == label + ":":
                ws.cell(row=row, column=2).value = field_value
        # Question rows: ID in col A, answer in col C
        if cell_text in answers:
            ws.cell(row=row, column=3).value = answers[cell_text]
    wb.save(path)


FULL_RESPONDENT = {"Name": "John Doe", "Institution": "Test University", "Email": "john@test.edu"}
FULL_ANSWERS = {"G.Q1": 4, "G.Q2": "Yes", "G.Q3": "Great!", "T.Q1": 7}


class TestValidResponse:
    def test_fully_filled_file_is_valid(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, FULL_ANSWERS)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert result.is_valid, result.errors

    def test_answers_are_extracted(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, FULL_ANSWERS)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert result.answers["G.Q1"] == 4
        assert result.answers["G.Q2"] == "Yes"
        assert result.answers["G.Q3"] == "Great!"
        assert result.answers["T.Q1"] == 7

    def test_respondent_info_is_extracted(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, FULL_ANSWERS)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert result.respondent_info["Name"] == "John Doe"
        assert result.respondent_info["Institution"] == "Test University"

    def test_config_hash_extracted(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, FULL_ANSWERS)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert result.config_hash == sample_questionnaire.config_hash()

    def test_optional_field_can_be_empty(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        respondent = {"Name": "Jane", "Institution": "Org"}  # Email omitted (optional)
        _fill_xlsx(generated_xlsx, respondent, FULL_ANSWERS)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert result.is_valid, result.errors

    def test_optional_question_can_be_empty(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        answers = {k: v for k, v in FULL_ANSWERS.items() if k != "G.Q3"}  # G.Q3 is optional
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, answers)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert result.is_valid, result.errors


class TestMissingSheets:
    def test_missing_questionnaire_sheet_fails(
        self, generated_xlsx: Path, sample_questionnaire, tmp_path: Path
    ) -> None:
        # Build a workbook that has _meta but no Questionnaire sheet.
        # (We cannot delete Questionnaire while keeping only the hidden _meta
        # because openpyxl refuses to save a workbook with no visible sheets.)
        src_wb = load_workbook(generated_xlsx, data_only=True)
        meta_rows = list(src_wb["_meta"].iter_rows(values_only=True))

        from openpyxl import Workbook as _WB
        new_wb = _WB()
        # Rename the default sheet to something other than Questionnaire
        new_wb.active.title = "Placeholder"
        ws_meta = new_wb.create_sheet("_meta")
        ws_meta.sheet_state = "hidden"
        for row in meta_rows:
            ws_meta.append(list(row))

        no_qs_path = tmp_path / "no_questionnaire.xlsx"
        new_wb.save(no_qs_path)

        result = validate_response(no_qs_path, sample_questionnaire)
        assert not result.is_valid
        assert any("Questionnaire" in e for e in result.errors)

    def test_missing_meta_sheet_fails(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        wb = load_workbook(generated_xlsx)
        del wb["_meta"]
        wb.save(generated_xlsx)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert not result.is_valid
        assert any("_meta" in e for e in result.errors)

    def test_corrupted_file_fails_gracefully(
        self, tmp_path: Path, sample_questionnaire
    ) -> None:
        bad_file = tmp_path / "corrupted.xlsx"
        bad_file.write_bytes(b"this is not a valid xlsx file at all")
        result = validate_response(bad_file, sample_questionnaire)
        assert not result.is_valid
        assert result.errors


class TestAnswerValidation:
    def test_missing_required_answer_fails(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, {})  # no answers
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert not result.is_valid
        assert any("required" in e.lower() for e in result.errors)

    def test_scale_value_below_min_fails(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        answers = {**FULL_ANSWERS, "G.Q1": 0}  # min is 1
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, answers)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert not result.is_valid
        assert any("G.Q1" in e and "minimum" in e for e in result.errors)

    def test_scale_value_above_max_fails(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        answers = {**FULL_ANSWERS, "G.Q1": 99}  # max is 5
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, answers)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert not result.is_valid
        assert any("G.Q1" in e and "maximum" in e for e in result.errors)

    def test_scale_boundary_values_are_valid(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        for boundary in (1, 5):
            _fill_xlsx(generated_xlsx, FULL_RESPONDENT, {**FULL_ANSWERS, "G.Q1": boundary})
            result = validate_response(generated_xlsx, sample_questionnaire)
            assert result.is_valid, f"Boundary value {boundary} should be valid"

    def test_scale_non_integer_fails(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        answers = {**FULL_ANSWERS, "G.Q1": "maybe"}
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, answers)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert not result.is_valid

    def test_yesno_invalid_value_fails(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        answers = {**FULL_ANSWERS, "G.Q2": "Maybe"}
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, answers)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert not result.is_valid
        assert any("G.Q2" in e for e in result.errors)

    @pytest.mark.parametrize("yesno_val", ["Yes", "No", "yes", "no", "YES", "NO"])
    def test_yesno_case_insensitive_valid(
        self, generated_xlsx: Path, sample_questionnaire, yesno_val: str
    ) -> None:
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, {**FULL_ANSWERS, "G.Q2": yesno_val})
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert result.is_valid, f"'{yesno_val}' should be valid for YES_NO"


class TestTampering:
    def test_tampered_question_ids_in_meta_fail(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        wb = load_workbook(generated_xlsx)
        ws_meta = wb["_meta"]
        for row in range(1, ws_meta.max_row + 1):
            if ws_meta.cell(row=row, column=1).value == "question_ids":
                ws_meta.cell(row=row, column=2).value = json.dumps(["FAKE.Q1", "FAKE.Q2"])
        wb.save(generated_xlsx)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert not result.is_valid

    def test_config_hash_mismatch_gives_warning_not_error(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        wb = load_workbook(generated_xlsx)
        ws_meta = wb["_meta"]
        for row in range(1, ws_meta.max_row + 1):
            if ws_meta.cell(row=row, column=1).value == "config_hash":
                ws_meta.cell(row=row, column=2).value = "aaaa" * 16  # fake hash
        wb.save(generated_xlsx)
        _fill_xlsx(generated_xlsx, FULL_RESPONDENT, FULL_ANSWERS)
        result = validate_response(generated_xlsx, sample_questionnaire)
        # Mismatch is a warning, not an error
        assert any("hash mismatch" in w.lower() for w in result.warnings)


class TestRespondentFieldValidation:
    def test_missing_required_respondent_field_fails(
        self, generated_xlsx: Path, sample_questionnaire
    ) -> None:
        respondent = {"Institution": "Org"}  # Name is required but missing
        _fill_xlsx(generated_xlsx, respondent, FULL_ANSWERS)
        result = validate_response(generated_xlsx, sample_questionnaire)
        assert not result.is_valid
        assert any("Name" in e for e in result.errors)


class TestI18nValidation:
    """Verify that yes/no validation uses language-specific strings."""

    def _make_german_xlsx(self, tmp_path, sample_questionnaire, sample_style):
        from umfrage.generator import generate_questionnaire
        q = sample_questionnaire.model_copy(update={"language": "de"})
        out = tmp_path / "de_questionnaire.xlsx"
        generate_questionnaire(q, sample_style, out)
        return out, q

    def test_german_ja_is_accepted(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        out, q = self._make_german_xlsx(tmp_path, sample_questionnaire, sample_style)
        _fill_xlsx(out, FULL_RESPONDENT, {**FULL_ANSWERS, "G.Q2": "Ja"})
        result = validate_response(out, q)
        assert result.is_valid, result.errors

    def test_german_nein_is_accepted(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        out, q = self._make_german_xlsx(tmp_path, sample_questionnaire, sample_style)
        _fill_xlsx(out, FULL_RESPONDENT, {**FULL_ANSWERS, "G.Q2": "Nein"})
        result = validate_response(out, q)
        assert result.is_valid, result.errors

    @pytest.mark.parametrize("german_val", ["ja", "JA", "Ja", "nein", "NEIN", "Nein"])
    def test_german_yesno_case_insensitive(
        self, tmp_path: Path, sample_questionnaire, sample_style, german_val: str
    ) -> None:
        out, q = self._make_german_xlsx(tmp_path, sample_questionnaire, sample_style)
        _fill_xlsx(out, FULL_RESPONDENT, {**FULL_ANSWERS, "G.Q2": german_val})
        result = validate_response(out, q)
        assert result.is_valid, f"'{german_val}' should be valid for German YES_NO"

    def test_german_questionnaire_rejects_english_yes(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        out, q = self._make_german_xlsx(tmp_path, sample_questionnaire, sample_style)
        _fill_xlsx(out, FULL_RESPONDENT, {**FULL_ANSWERS, "G.Q2": "Yes"})
        result = validate_response(out, q)
        assert not result.is_valid
        assert any("G.Q2" in e for e in result.errors)

    def test_german_questionnaire_rejects_english_no(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        out, q = self._make_german_xlsx(tmp_path, sample_questionnaire, sample_style)
        _fill_xlsx(out, FULL_RESPONDENT, {**FULL_ANSWERS, "G.Q2": "No"})
        result = validate_response(out, q)
        assert not result.is_valid
